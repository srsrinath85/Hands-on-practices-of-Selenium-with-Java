# import xlrd
import openpyxl
# excel data accessing
# jg=openpyxl.load_workbook("C://Users//Admin//Documents//datafiles.xlsx")
# fj=jg['dataset']
# kj=fj.max_row
# kl=fj.max_column
# print(kj)
# print(kl)
# for crow in range(1,kj+1):
#     for col in range(1,kl+1):
#         print(fj.cell(crow,col).value)

# wb=openpyxl.load_workbook("C://Users//Admin//Documents//datafiles.xlsx")
# ws=wb['dataset']
# rowcount=ws.max_row
# colcount=ws.max_column
# print("total number of rows are:",+rowcount)
# print("total number of columns are:",+colcount)
# print('The value in cell A1 is: '+ws['A1'].value)

# for i in range(1,ws.max_column+1):
#     values=ws.cell(row=1,column=i).value
#     print(values)

# for i in range(1,ws.max_row+1):
#     valuess=ws.cell(row=i,column=1).value
#     value=ws.cell(row=i,column=2).value
#     print(valuess)
#     print(value)

# for i in range(2,4):
#     valuess=ws.cell(row=i,column=1).value
#     valu=ws.cell(row=i,column=2).value
#     print(valuess)
#     print(valu)

#Getting excel data in a table formate using iter_row method in python and list 

# my_list=list()

# for val in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=4,values_only=True):
#     my_list.append(val)

# for ele1,ele2,ele3,ele4 in my_list:
#     (print("{:<10}{:<13}{:<10}{:<15}".format(ele1,ele2,ele3,ele4)))




#writing data in excel file in row and column

wb=openpyxl.load_workbook("C://Users//Admin//Documents//datafiles.xlsx")
ws=wb['dataset']

# ws['E1']='Name of the person'
# wb.save("C://Users//Admin//Documents//datafiles.xlsx")
# row_position=2
# col_position='E1'
# total_values=(ws.cell(row=row_position,column=col_position).)+
# wd=ws.cell(row=)
# ws.cell(row=7,column=5,value='saketh')
# ws.cell(row=8,column=5,value='somu')
# ws.cell(row=9,column=5,value='sasi')
# ws.cell(row=10,column=5,value='shashank')
# ws.cell(row=11,column=5,value='shashank kumar')
# wb.save("C://Users//Admin//Documents//datafiles.xlsx")