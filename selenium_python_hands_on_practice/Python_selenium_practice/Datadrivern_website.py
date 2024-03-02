from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl


driver=webdriver.Chrome()
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
driver.maximize_window()
time.sleep(5)
wb=openpyxl.load_workbook("C://Users//Admin//Documents//site_name.xlsx")
ws=wb['Dat']
roc=ws.max_row
clc=ws.max_column

for d in range(2,ws.max_row+1):
        val=ws.cell(row=d,column=1).value
        val1=ws.cell(row=d,column=2).value
        print(val)
        print(val1)
        driver.find_element(By.NAME,'username').send_keys(val)
        driver.find_element(By.NAME,'password').send_keys(val1)
        driver.find_element(By.XPATH,"//*[@type='submit']").click()
        time.sleep(5)
        sr=driver.current_url   
        print(sr)        
        if sr=='https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index':
            print('opening new URL')
            for t in range(3,ws.max_row+1):
                val4=ws.cell(row=t,column=1).value
                val2=ws.cell(row=t,column=2).value
                print(val4)
                print(val2)
                driver.get("https://practicetestautomation.com/practice-test-login/")
                driver.find_element(By.NAME,'username').send_keys(val4)
                driver.find_element(By.NAME,'password').send_keys(val2)
                time.sleep(5)
                driver.find_element(By.ID,'submit').click()
                time.sleep(6)
